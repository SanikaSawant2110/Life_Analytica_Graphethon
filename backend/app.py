from flask import Flask, jsonify, request
from flask_cors import CORS
import json, random, math, os
from datetime import datetime, timedelta
from google import genai
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter  # kept for append_to_sensor_log column formatting

app = Flask(__name__)
CORS(app)

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
DATA_DIR       = os.path.join(BASE_DIR, '..', 'data')
SENSOR_FILE     = os.path.abspath(os.path.join(DATA_DIR, 'LifeAnalytica_SensorLog.xlsx'))
TRANSCRIPT_FILE = os.path.abspath(os.path.join(DATA_DIR, 'VoiceTranscripts.xlsx'))  # write-only archive

thin = Side(style="thin", color="E2E8F0")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── Generate 120 minutes of realistic sensor data ───────────────────────────
def generate_sensor_data():
    rows = []
    t = datetime(2025, 4, 4, 9, 0, 0)
    hr = 72.0; spo2 = 98.0; temp = 36.5

    def zone(i):
        if   i <  15: return "rest"
        elif i <  30: return "walk"
        elif i <  45: return "stress"
        elif i <  60: return "rest"
        elif i <  75: return "exercise"
        elif i <  90: return "recovery"
        elif i < 105: return "conversation"
        else:         return "rest"

    targets = {
        "rest":         (68,  98.5, 0.02, 0.03),
        "walk":         (82,  98.0, 0.60, 0.80),
        "stress":       (98,  97.2, 0.30, 0.30),
        "exercise":     (118, 96.8, 1.20, 1.50),
        "recovery":     (88,  97.5, 0.20, 0.20),
        "conversation": (78,  98.0, 0.15, 0.10),
    }
    audio_targets = {
        "stress":       (65, 250),
        "conversation": (55, 200),
        "exercise":     (60, 120),
        "walk":         (45,  80),
        "rest":         (35,  60),
        "recovery":     (38,  65),
    }

    for i in range(120):
        z = zone(i)
        n = lambda s: random.gauss(0, s)
        hr_t, sp_t, ax_t, ay_t = targets.get(z, (70, 98.2, 0.02, 0.03))
        db_t, freq_t = audio_targets.get(z, (35, 60))

        hr   = max(50,  min(150, hr   * 0.85 + hr_t  * 0.15 + n(2)))
        spo2 = max(94,  min(100, spo2 * 0.90 + sp_t  * 0.10 + n(0.2)))
        temp = max(35.5,min(37.8, temp + n(0.04)))
        ax   = ax_t  + n(0.05);  ay = ay_t + n(0.05); az = 9.81 + n(0.08)
        gx   = n(0.4); gy = n(0.4); gz = n(0.2)
        db   = max(20, db_t   + n(4))
        freq = max(20, freq_t + n(15))

        flag = "normal"
        if hr > 100:   flag = "high_hr"
        if spo2 < 97:  flag = "low_spo2"
        if z == "stress" and hr > 90: flag = "stress"

        rows.append({
            "timestamp": t.strftime("%H:%M"),
            "minute":    i,
            "zone":      z,
            "hr":        round(hr),
            "spo2":      round(spo2, 1),
            "temp":      round(temp, 2),
            "ax":        round(ax, 3),
            "ay":        round(ay, 3),
            "az":        round(az, 3),
            "gx":        round(gx, 3),
            "gy":        round(gy, 3),
            "gz":        round(gz, 3),
            "db":        round(db, 1),
            "freq":      round(freq, 1),
            "flag":      flag,
            "voice_transcript": None,
        })
        t += timedelta(minutes=1)
    return rows

SENSOR_DATA = generate_sensor_data()

SEMANTIC_SEGMENTS = [
    {"window": "09:00–09:15", "zone": "rest",         "avg_hr": 70,  "observation": "Baseline state. Low HR, normal SpO2, minimal movement.", "note": "Patient appears calm at session start."},
    {"window": "09:15–09:30", "zone": "walk",         "avg_hr": 82,  "observation": "HR rose ~15 bpm. Elevated accelerometer Y. Rhythmic motion detected.", "note": "Light physical activity — possible commute or warmup walk."},
    {"window": "09:30–09:45", "zone": "stress",       "avg_hr": 98,  "observation": "Sharp HR spike to 98+. Audio dB >65. SpO2 dipped to 97.2%. Gyro elevated.", "note": "⚠ Stress/conflict episode. Raised voice detected in audio. Clinically significant."},
    {"window": "09:45–10:00", "zone": "rest",         "avg_hr": 70,  "observation": "HR normalised. Temp stable. Low motion and sound levels.", "note": "Recovery after stress episode. Monitor for recurrence patterns."},
    {"window": "10:00–10:15", "zone": "exercise",     "avg_hr": 118, "observation": "HR peak ~118. High accel X/Y values. Audio ~60 dB. Vigorous movement.", "note": "Intentional physical activity. Expected HR elevation — not pathological."},
    {"window": "10:15–10:30", "zone": "recovery",     "avg_hr": 88,  "observation": "Gradual HR drop post-exercise. SpO2 recovering to 97.5%. Reduced movement.", "note": "Normal post-exercise recovery trajectory. Good cardiovascular response."},
    {"window": "10:30–10:45", "zone": "conversation", "avg_hr": 78,  "observation": "Moderate HR. Audio dominant frequency ~200 Hz (voice). Low accel.", "note": "Social interaction detected. Calm tone inferred from audio features."},
    {"window": "10:45–11:00", "zone": "rest",         "avg_hr": 70,  "observation": "Baseline restored. All vitals within normal range. Low movement.", "note": "Session ending in calm state — positive prognostic sign."},
]

# ─── Helper: generate random biometrics for a voice session ──────────────────
def generate_voice_biometrics(zone: str) -> dict:
    """Generate realistic random sensor readings for a voice recording moment."""
    targets = {
        "stress":       {"hr": (95, 115), "spo2": (96.5, 97.8), "temp": (36.4, 36.9), "db": (62, 75), "freq": (220, 280), "ax": (0.25, 0.50), "ay": (0.20, 0.45)},
        "rest":         {"hr": (62, 78),  "spo2": (97.8, 99.2), "temp": (36.2, 36.6), "db": (28, 42), "freq": (45, 80),   "ax": (0.01, 0.06), "ay": (0.01, 0.06)},
        "exercise":     {"hr": (105, 130),"spo2": (96.0, 97.5), "temp": (36.6, 37.2), "db": (55, 68), "freq": (100, 140), "ax": (0.80, 1.60), "ay": (0.80, 1.60)},
        "conversation": {"hr": (72, 88),  "spo2": (97.5, 99.0), "temp": (36.3, 36.7), "db": (48, 62), "freq": (180, 230), "ax": (0.08, 0.20), "ay": (0.05, 0.15)},
        "recovery":     {"hr": (80, 95),  "spo2": (97.0, 98.5), "temp": (36.4, 36.8), "db": (32, 48), "freq": (60, 100),  "ax": (0.10, 0.25), "ay": (0.08, 0.20)},
        "walk":         {"hr": (78, 95),  "spo2": (97.2, 98.8), "temp": (36.3, 36.7), "db": (40, 52), "freq": (70, 110),  "ax": (0.50, 0.90), "ay": (0.50, 0.90)},
    }
    t = targets.get(zone, targets["rest"])
    def rr(lo, hi, dec=1): return round(random.uniform(lo, hi), dec)
    return {
        "hr":    rr(*t["hr"], dec=0),
        "spo2":  rr(*t["spo2"], dec=1),
        "temp":  rr(*t["temp"], dec=2),
        "db":    rr(*t["db"], dec=1),
        "freq":  rr(*t["freq"], dec=1),
        "ax":    rr(*t["ax"], dec=3),
        "ay":    rr(*t["ay"], dec=3),
        "az":    round(9.81 + random.gauss(0, 0.06), 3),
        "gx":    round(random.gauss(0, 0.4), 3),
        "gy":    round(random.gauss(0, 0.4), 3),
    }

# ─── Helper: ensure SensorLog has transcript column ──────────────────────────
def ensure_sensor_log_transcript_col():
    """Make sure column N (14) has the Voice Transcript header in SensorLog."""
    if not os.path.exists(SENSOR_FILE):
        return
    wb = load_workbook(SENSOR_FILE)
    ws = wb['Sensor Log']
    # Check if col 14 header already set
    if ws.cell(3, 14).value == 'Transcript':
        wb.close()
        return
    # Add headers
    ws.cell(2, 14).value = 'Voice Transcript'
    ws.cell(2, 14).font  = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
    ws.cell(2, 14).fill  = PatternFill('solid', start_color='6D28D9')
    ws.cell(2, 14).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(2, 14).border = brd

    ws.cell(3, 14).value = 'Transcript'
    ws.cell(3, 14).font  = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
    ws.cell(3, 14).fill  = PatternFill('solid', start_color='6D28D9')
    ws.cell(3, 14).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(3, 14).border = brd

    ws.column_dimensions['N'].width = 60
    wb.save(SENSOR_FILE)

# ─── Helper: append a new row to SensorLog ───────────────────────────────────
ZONE_COLORS_XL = {
    "stress":       {"row": "FFF0F0", "flag_color": "CC0000"},
    "rest":         {"row": "F0F7FF", "flag_color": "1A6BCC"},
    "exercise":     {"row": "FFFBF0", "flag_color": "D97706"},
    "conversation": {"row": "F5F0FF", "flag_color": "6D28D9"},
    "recovery":     {"row": "F0FFF4", "flag_color": "16A34A"},
    "walk":         {"row": "F0FFF0", "flag_color": "16A34A"},
}

def append_to_sensor_log(timestamp: str, zone: str, bio: dict, transcript: str, flag: str) -> int:
    """Append one row to the Sensor Log sheet and return the Excel row number."""
    ensure_sensor_log_transcript_col()
    wb = load_workbook(SENSOR_FILE)
    ws = wb['Sensor Log']

    next_row = ws.max_row + 1
    zc = ZONE_COLORS_XL.get(zone, {"row": "FFFFFF", "flag_color": "000000"})
    bg = PatternFill('solid', start_color=zc["row"])

    flag_text = f"⚠ {flag.replace('_',' ').title()}" if flag != "normal" else "✓ Normal"
    flag_color = zc["flag_color"] if flag != "normal" else "16A34A"

    values = [
        timestamp,          # A - Time
        bio["db"],          # B - dB
        bio["freq"],        # C - Freq(Hz)
        bio["hr"],          # D - HR(bpm)
        bio["spo2"],        # E - SpO2(%)
        bio["temp"],        # F - Temp(°C)
        bio["ax"],          # G - Ax(g)
        bio["ay"],          # H - Ay(g)
        bio["az"],          # I - Az(g)
        bio["gx"],          # J - Gx(°/s)
        bio["gy"],          # K - Gy(°/s)
        zone.title(),       # L - Zone
        flag_text,          # M - Flag
        transcript,         # N - Voice Transcript
    ]

    for ci, val in enumerate(values, 1):
        cell = ws.cell(next_row, ci)
        cell.value = val
        cell.fill  = bg
        cell.border = brd

        is_flag_col = (ci == 13)
        is_transcript_col = (ci == 14)
        cell.font = Font(
            name='Calibri', size=9,
            color=flag_color if is_flag_col else ("6D28D9" if is_transcript_col and transcript else "000000"),
            bold=(is_flag_col and flag != "normal"),
            italic=(is_transcript_col and bool(transcript)),
        )
        cell.alignment = Alignment(
            horizontal='left' if is_transcript_col else 'center',
            vertical='center',
            wrap_text=is_transcript_col,
        )

    if transcript:
        ws.row_dimensions[next_row].height = max(30, min(120, len(transcript) // 3))
    else:
        ws.row_dimensions[next_row].height = 18

    wb.save(SENSOR_FILE)
    return next_row

# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route("/api/sensor-log")
def sensor_log():
    return jsonify(SENSOR_DATA)

@app.route("/api/semantic-timeline")
def semantic_timeline():
    return jsonify(SEMANTIC_SEGMENTS)

@app.route("/api/hourly-summary")
def hourly_summary():
    h1 = [r for r in SENSOR_DATA if r["minute"] < 60]
    h2 = [r for r in SENSOR_DATA if r["minute"] >= 60]
    def summarise(rows, label):
        hrs   = [r["hr"]   for r in rows]
        spo2s = [r["spo2"] for r in rows]
        temps = [r["temp"] for r in rows]
        accs  = [math.sqrt(r["ax"]**2 + r["ay"]**2) for r in rows]
        dbs   = [r["db"]   for r in rows]
        zones = {}
        for r in rows:
            zones[r["zone"]] = zones.get(r["zone"], 0) + 1
        dominant_zone = max(zones, key=zones.get)
        flag = "normal"
        if max(hrs) > 100: flag = "elevated_hr"
        if min(spo2s) < 97: flag = "low_spo2"
        return {
            "label":         label,
            "avg_hr":        round(sum(hrs)/len(hrs)),
            "max_hr":        max(hrs),
            "min_hr":        min(hrs),
            "avg_spo2":      round(sum(spo2s)/len(spo2s), 1),
            "min_spo2":      min(spo2s),
            "avg_temp":      round(sum(temps)/len(temps), 2),
            "avg_acc":       round(sum(accs)/len(accs), 3),
            "avg_db":        round(sum(dbs)/len(dbs), 1),
            "dominant_zone": dominant_zone,
            "zone_breakdown": zones,
            "flag":          flag,
        }
    return jsonify([summarise(h1, "09:00–10:00"), summarise(h2, "10:00–11:00")])

@app.route("/api/vitals-chart")
def vitals_chart():
    sampled = [r for r in SENSOR_DATA if r["minute"] % 5 == 0]
    return jsonify([{
        "timestamp": r["timestamp"],
        "hr":        r["hr"],
        "spo2":      r["spo2"],
        "temp":      r["temp"],
        "db":        r["db"],
        "ax":        round(math.sqrt(r["ax"]**2 + r["ay"]**2), 3),
        "zone":      r["zone"],
    } for r in sampled])

@app.route("/api/ai-summary", methods=["POST"])
def ai_summary():
    if not GEMINI_API_KEY:
        return jsonify({"summary": mock_clinical_summary(), "source": "mock"})
    try:
        data = request.json or {}
        stats = data.get("stats", build_stats_for_gemini())
        client = genai.Client(api_key=GEMINI_API_KEY)
        prompt = f"""
You are a clinical AI assistant analyzing wearable sensor data for a therapist or psychiatrist.

Patient session data (09:00–11:00):
{json.dumps(stats, indent=2)}

Generate a structured clinical summary with:
1. OVERALL STATE (1 sentence)
2. KEY FINDINGS (3-4 bullet points with clinical language)
3. STRESS INDICATORS (what the data reveals)
4. THERAPIST RECOMMENDATIONS (2-3 practical, evidence-based suggestions)
5. RISK FLAGS (if any, or "None identified")

Keep it concise, clinically appropriate, and evidence-based. Use medical terminology where appropriate.
Format with clear section headers.
"""
        response = client.models.generate_content(model="gemini-2.0-flash", contents=prompt)
        return jsonify({"summary": response.text, "source": "gemini"})
    except Exception as e:
        return jsonify({"summary": mock_clinical_summary(), "source": "mock", "error": str(e)})

def build_stats_for_gemini():
    hrs   = [r["hr"]   for r in SENSOR_DATA]
    spo2s = [r["spo2"] for r in SENSOR_DATA]
    temps = [r["temp"] for r in SENSOR_DATA]
    return {
        "session_duration_minutes": 120,
        "avg_heart_rate":  round(sum(hrs)/len(hrs)),
        "max_heart_rate":  max(hrs),
        "min_heart_rate":  min(hrs),
        "avg_spo2":        round(sum(spo2s)/len(spo2s), 1),
        "min_spo2":        min(spo2s),
        "avg_temp":        round(sum(temps)/len(temps), 2),
        "stress_minutes":  sum(1 for r in SENSOR_DATA if r["zone"] == "stress"),
        "exercise_minutes":sum(1 for r in SENSOR_DATA if r["zone"] == "exercise"),
        "rest_minutes":    sum(1 for r in SENSOR_DATA if r["zone"] == "rest"),
        "high_audio_events":   sum(1 for r in SENSOR_DATA if r["db"] > 60),
        "hr_spikes_over_100":  sum(1 for r in SENSOR_DATA if r["hr"] > 100),
        "low_spo2_events":     sum(1 for r in SENSOR_DATA if r["spo2"] < 97),
    }

def mock_clinical_summary():
    return """## OVERALL STATE
Patient demonstrated a variable physiological pattern across the 2-hour session, with a notable stress episode at 09:30–09:45 and healthy physical activity response.

## KEY FINDINGS
- **Heart Rate**: Average 82 bpm (range: 68–118 bpm). Elevated HR observed during exercise (expected) and one anomalous stress spike at 09:35.
- **SpO2**: Average 97.8%. Minor dip to 97.2% during stress episode — clinically within acceptable range but warrants monitoring.
- **Skin Temperature**: Stable at 36.5°C average. No fever indicators.
- **Activity**: Moderate physical activity with clear exercise bout (10:00–10:15) showing healthy cardiovascular response and normal recovery curve.

## STRESS INDICATORS
The sensor fusion reveals a **15-minute stress episode (09:30–09:45)** characterised by: concurrent HR elevation (+26 bpm above baseline), elevated ambient audio (dB >65 with voice-frequency dominance at 250 Hz), and increased gyroscopic variance suggesting restlessness. This pattern is consistent with an acute psychosocial stressor — possibly conflict or high-pressure conversation.

## THERAPIST RECOMMENDATIONS
- **Trigger Identification**: Explore what occurred at 09:30. Audio frequency pattern suggests verbal conflict or heightened emotional exchange.
- **Stress Response Monitoring**: The rapid HR normalisation (within 15 min) suggests adequate stress recovery capacity — reinforce this as a resilience indicator.
- **Physical Activity as Regulation**: The exercise bout at 10:00 may have been a self-regulatory response. Explore if patient uses physical activity consciously for mood regulation.

## RISK FLAGS
- ⚠ Single stress spike noted (09:30–09:45) — monitor for pattern recurrence across sessions
- SpO2 dip to 97.2% during stress — benign in isolation, flag if recurrent"""

@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "version": "2.0.0", "service": "LifeAnalytica Clinical API"})

def detect_emotion_zone(text: str) -> str:
    text_lower = text.lower()
    stress_kw    = ['anxious','panic','stressed','overwhelmed','scared','worried','fear','afraid','nervous','tense','angry','frustrated',"can't breathe",'heart racing','pressure','terrified','dread']
    exercise_kw  = ['running','jogging','workout','gym','exercise','walking','tired','exhausted','breathing hard']
    calm_kw      = ['calm','relaxed','fine','good','okay','happy','peaceful','better','comfortable']
    convo_kw     = ['talking','conversation','discussing','meeting','call','spoke','said','told']
    scores = {
        'stress':       sum(1 for kw in stress_kw   if kw in text_lower),
        'exercise':     sum(1 for kw in exercise_kw if kw in text_lower),
        'conversation': sum(1 for kw in convo_kw    if kw in text_lower),
        'rest':         sum(1 for kw in calm_kw     if kw in text_lower),
    }
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else 'rest'

def extract_keywords(text: str) -> str:
    stop = {'i','me','my','the','a','an','is','was','and','or','but','in','on','at','to','for','of','it','this','that','with','have','had','be','been','are','were','so','do','did','not','no','yes'}
    words = [w.strip('.,!?;:"\'').lower() for w in text.split()]
    kws   = [w for w in words if len(w) > 3 and w not in stop]
    seen = set(); unique = []
    for k in kws:
        if k not in seen: seen.add(k); unique.append(k)
    return ', '.join(unique[:6])

def _read_sensor_transcripts(zone_filter=None):
    """
    Read rows from LifeAnalytica_SensorLog.xlsx that have a transcript (col N).
    SensorLog columns (0-indexed): 0=Time, 1=dB, 2=Freq, 3=HR, 4=SpO2, 5=Temp,
    6=Ax, 7=Ay, 8=Az, 9=Gx, 10=Gy, 11=Zone, 12=Flag, 13=Transcript
    zone_filter: if set, only return rows where zone matches (case-insensitive)
    """
    if not os.path.exists(SENSOR_FILE):
        return []
    wb = load_workbook(SENSOR_FILE, data_only=True)
    ws = wb['Sensor Log']
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):  # skip 3 header rows
        transcript = row[13]
        if not transcript:
            continue
        zone = str(row[11] or '').lower()
        if zone_filter and zone != zone_filter.lower():
            continue
        flag_raw = str(row[12] or '').lower()
        flag = 'stress' if 'stress' in flag_raw else 'normal'
        keywords = extract_keywords(str(transcript))
        rows.append({
            "timestamp":  str(row[0] or ''),
            "duration":   0,
            "zone":       zone,
            "hr":         row[3] if isinstance(row[3], (int, float)) else None,
            "spo2":       row[4] if isinstance(row[4], (int, float)) else None,
            "transcript": str(transcript),
            "keywords":   keywords,
            "flag":       flag,
        })
    return rows

@app.route("/api/transcripts", methods=["GET"])
def get_transcripts():
    """Return ALL rows from SensorLog that have a transcript."""
    try:
        return jsonify(_read_sensor_transcripts())
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _archive_to_voice_transcripts(timestamp, duration, zone, bio, transcript, keywords, flag):
    """Write-only archive to VoiceTranscripts.xlsx. Never read back by the app."""
    zone_colors = {"rest": "D6EAF8", "stress": "FDECEA", "exercise": "FFF3CD", "conversation": "F5EEF8", "recovery": "EAF0FB"}
    bg = zone_colors.get(zone, "FFFFFF")

    if not os.path.exists(TRANSCRIPT_FILE):
        os.makedirs(os.path.dirname(TRANSCRIPT_FILE), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Voice Log"
        ws.merge_cells("A1:H1")
        ws["A1"].value = "LifeAnalytica — Voice Transcript Archive (write-only)"
        ws["A1"].font  = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
        ws["A1"].fill  = PatternFill("solid", start_color="001F5B")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        for ci, h in enumerate(["Timestamp","Duration (s)","Zone","HR (bpm)","SpO2 (%)","Transcript","Keywords","Flag"], 1):
            c = ws.cell(2, ci)
            c.value = h
            c.font  = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            c.fill  = PatternFill("solid", start_color="1A6BCC")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd
        for ci, w in enumerate([20,12,16,12,12,70,35,14], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A3"
        wb.save(TRANSCRIPT_FILE)

    wb = load_workbook(TRANSCRIPT_FILE)
    ws = wb["Voice Log"]
    nr = ws.max_row + 1
    for ci, v in enumerate([timestamp, round(duration,1), zone.title(), bio["hr"], bio["spo2"], transcript, keywords, flag], 1):
        c = ws.cell(nr, ci)
        c.value = v
        c.font  = Font(size=10, name="Calibri",
                       bold=(ci==8 and flag=="stress"),
                       color="CC0000" if (ci==8 and flag=="stress") else "000000")
        c.fill  = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left" if ci==6 else "center", vertical="center", wrap_text=(ci==6))
        c.border = brd
    ws.row_dimensions[nr].height = max(30, len(transcript)//4)
    wb.save(TRANSCRIPT_FILE)

@app.route("/api/transcripts", methods=["POST"])
def save_transcript():
    """Save voice transcript → SensorLog (source of truth) + VoiceTranscripts.xlsx (write-only archive)."""
    data       = request.json or {}
    transcript = data.get("transcript", "").strip()
    duration   = data.get("duration", 0)
    timestamp  = data.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    if not transcript:
        return jsonify({"error": "Empty transcript"}), 400

    zone     = detect_emotion_zone(transcript)
    keywords = extract_keywords(transcript)
    flag     = "stress" if zone == "stress" else "normal"
    bio      = generate_voice_biometrics(zone)

    # ── 1. Primary: append to LifeAnalytica_SensorLog.xlsx ────────────────────
    sensor_row = None
    sensor_error = None
    try:
        sensor_row = append_to_sensor_log(timestamp, zone, bio, transcript, flag)
    except Exception as e:
        sensor_error = str(e)

    # ── 2. Archive-only: write to VoiceTranscripts.xlsx (never read back) ─────
    try:
        _archive_to_voice_transcripts(timestamp, duration, zone, bio, transcript, keywords, flag)
    except Exception:
        pass  # Archive failure is silent — SensorLog is the source of truth

    return jsonify({
        "success":      True,
        "timestamp":    timestamp,
        "zone":         zone,
        "keywords":     keywords,
        "flag":         flag,
        "biometrics":   bio,
        "sensor_row":   sensor_row,
        "sensor_error": sensor_error,
    })

@app.route("/api/transcripts/stress", methods=["GET"])
def get_stress_transcripts():
    """Return only stress-flagged transcript rows from SensorLog."""
    try:
        return jsonify(_read_sensor_transcripts(zone_filter='stress'))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/transcripts/export-path", methods=["GET"])
def transcript_export_path():
    return jsonify({
        "sensor_log_path":   SENSOR_FILE,
        "sensor_log_exists": os.path.exists(SENSOR_FILE),
    })

if __name__ == "__main__":
    ensure_sensor_log_transcript_col()
    app.run(debug=True, port=5050)